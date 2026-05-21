from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.models.api import WeeklyDigest
from app.services.report_generator import ReportExcelGenerator, ReportPDFGenerator


def _mock_digest() -> WeeklyDigest:
    backend_root = Path(__file__).resolve().parents[2]
    payload = json.loads(
        (backend_root / "docs/api/mock/weekly-digest.detail.response.json").read_text()
    )
    return WeeklyDigest.model_validate(payload)


def test_report_generators_create_valid_files() -> None:
    digest = _mock_digest()

    pdf = ReportPDFGenerator(digest).generate()
    xlsx = ReportExcelGenerator(digest).generate()

    assert pdf.startswith(b"%PDF-")
    workbook = load_workbook(BytesIO(xlsx))
    assert workbook.sheetnames == ["Summary", "Analytics", "Trackers", "Threats"]


def test_excel_report_uses_readable_enterprise_formatting() -> None:
    digest = _mock_digest()

    workbook = load_workbook(BytesIO(ReportExcelGenerator(digest).generate()))
    summary = workbook["Summary"]
    analytics = workbook["Analytics"]
    trackers = workbook["Trackers"]
    threats = workbook["Threats"]

    assert summary["A1"].value == "Weekly Digest Report"
    assert summary["A14"].value == "Event Signals"
    assert summary.column_dimensions["A"].width >= 20
    assert analytics["A3"].value == "Event Type Distribution"
    assert len(analytics._charts) == 3
    assert trackers["B2"].value == "CATEGORY"
    assert "TrackerType." not in trackers["B2"].value
    assert threats["E1"].value == "Trackers"
    assert threats.column_dimensions["C"].width >= 50
    assert threats["D2"].alignment.wrap_text is True
