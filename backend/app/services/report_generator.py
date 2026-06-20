from __future__ import annotations

from collections import Counter
from copy import copy
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.api import WeeklyDigest


BRAND_NAVY = "17324D"
BRAND_BLUE = "2563EB"
INK = "111827"
MUTED = "6B7280"
SURFACE = "F8FAFC"
BORDER = "D9E2EC"
HEADER_FILL = "EAF1F8"
WARN_FILL = "FFF7ED"
GREEN = "059669"
AMBER = "D97706"


def _value(value: Any) -> str:
    enum_value = getattr(value, "value", value)
    return "" if enum_value is None else str(enum_value)


def _join_values(values: list[Any] | None) -> str:
    return ", ".join(_value(value) for value in values or [])


def _safe_pdf_text(value: Any) -> str:
    text = _value(value)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _pdf_rgb(hex_color: str) -> tuple[int, int, int]:
    return (
        int(hex_color[0:2], 16),
        int(hex_color[2:4], 16),
        int(hex_color[4:6], 16),
    )


@dataclass(frozen=True)
class ReportAnalytics:
    event_type_counts: list[tuple[str, int]]
    marketplace_counts: list[tuple[str, int]]
    tracker_threat_counts: list[tuple[str, int]]
    total_event_signals: int
    average_signals_per_threat: float
    active_marketplaces: int
    top_event_type: str
    top_tracker: str


def _sorted_counts(counter: Counter[str]) -> list[tuple[str, int]]:
    return sorted(counter.items(), key=lambda item: (-item[1], item[0]))


def _build_analytics(digest: WeeklyDigest) -> ReportAnalytics:
    event_type_counter: Counter[str] = Counter()
    marketplace_counter: Counter[str] = Counter()
    tracker_counter: Counter[str] = Counter()

    for threat in digest.threats:
        marketplace_counter[_value(threat.marketplace)] += 1
        event_type_counter.update(_value(event_type) for event_type in threat.event_types or [])
        tracker_counter.update(ref.tracker_name for ref in threat.tracker_refs)

    event_type_counts = _sorted_counts(event_type_counter)
    tracker_threat_counts = _sorted_counts(tracker_counter)
    total_event_signals = sum(event_type_counter.values())
    threat_count = len(digest.threats)

    return ReportAnalytics(
        event_type_counts=event_type_counts,
        marketplace_counts=_sorted_counts(marketplace_counter),
        tracker_threat_counts=tracker_threat_counts,
        total_event_signals=total_event_signals,
        average_signals_per_threat=(
            round(total_event_signals / threat_count, 1) if threat_count else 0.0
        ),
        active_marketplaces=len(marketplace_counter),
        top_event_type=event_type_counts[0][0] if event_type_counts else "N/A",
        top_tracker=tracker_threat_counts[0][0] if tracker_threat_counts else "N/A",
    )


class ReportPDFGenerator:
    def __init__(self, digest: "WeeklyDigest") -> None:
        self.digest = digest
        self.analytics = _build_analytics(digest)

    def generate(self) -> bytes:
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_margins(14, 14, 14)
        pdf.set_auto_page_break(auto=True, margin=14)
        pdf.add_page()

        self._draw_header(pdf)
        self._draw_summary(pdf)
        self._draw_signal_analytics(pdf)
        if self.digest.insights:
            self._draw_ai_insights(pdf)
        self._draw_trackers(pdf)
        self._draw_threats(pdf)

        return bytes(pdf.output())

    def _set_text_color(self, pdf: FPDF, hex_color: str) -> None:
        pdf.set_text_color(*_pdf_rgb(hex_color))

    def _set_fill_color(self, pdf: FPDF, hex_color: str) -> None:
        pdf.set_fill_color(*_pdf_rgb(hex_color))

    def _section_title(self, pdf: FPDF, title: str) -> None:
        pdf.ln(5)
        self._set_text_color(pdf, BRAND_NAVY)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self._set_text_color(pdf, INK)

    def _draw_header(self, pdf: FPDF) -> None:
        page_width = pdf.w - pdf.l_margin - pdf.r_margin
        self._set_fill_color(pdf, BRAND_NAVY)
        pdf.rect(pdf.l_margin, 12, page_width, 28, "F")

        pdf.set_xy(pdf.l_margin + 5, 17)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 17)
        pdf.cell(
            page_width - 10,
            8,
            "Weekly Digest Report",
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )

        pdf.set_x(pdf.l_margin + 5)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(
            page_width - 10,
            6,
            _safe_pdf_text(f"{self.digest.week_start} to {self.digest.week_end}"),
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )

        pdf.set_y(46)
        self._set_text_color(pdf, INK)
        self._set_fill_color(pdf, SURFACE)
        pdf.rect(pdf.l_margin, pdf.get_y(), page_width, 20, "F")
        pdf.set_draw_color(*_pdf_rgb(BORDER))
        pdf.rect(pdf.l_margin, pdf.get_y(), page_width, 20)

        meta = [
            ("Digest Code", self.digest.digest_code),
            ("Created", self.digest.created_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Trackers", len(self.digest.tracker_refs)),
        ]
        col_width = page_width / 3
        y = pdf.get_y() + 4
        for index, (label, value) in enumerate(meta):
            x = pdf.l_margin + (index * col_width) + 5
            pdf.set_xy(x, y)
            self._set_text_color(pdf, MUTED)
            pdf.set_font("Helvetica", "B", 7)
            pdf.cell(
                col_width - 10,
                4,
                label.upper(),
                new_x=XPos.LMARGIN,
                new_y=YPos.NEXT,
            )
            pdf.set_x(x)
            self._set_text_color(pdf, INK)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(col_width - 10, 5, _safe_pdf_text(value))
        pdf.set_y(72)

    def _draw_summary(self, pdf: FPDF) -> None:
        self._section_title(pdf, "Executive Summary")
        metrics = [
            ("New Entrants", self.digest.summary.new_entrant_count),
            ("Returning", self.digest.summary.returning_count),
            ("Top 10 Entry", self.digest.summary.top10_enter_count),
            ("Price Changes", self.digest.summary.price_change_count),
            ("Listing Changes", self.digest.summary.listing_change_count),
            ("Threats", len(self.digest.threats)),
            ("Event Signals", self.analytics.total_event_signals),
            ("Marketplaces", self.analytics.active_marketplaces),
        ]
        page_width = pdf.w - pdf.l_margin - pdf.r_margin
        gap = 3
        columns = 4
        card_width = (page_width - gap * (columns - 1)) / columns
        card_height = 20
        y = pdf.get_y()
        for index, (label, value) in enumerate(metrics):
            row = index // columns
            column = index % columns
            x = pdf.l_margin + column * (card_width + gap)
            card_y = y + row * (card_height + gap)
            self._set_fill_color(pdf, SURFACE)
            pdf.set_draw_color(*_pdf_rgb(BORDER))
            pdf.rect(x, card_y, card_width, card_height, "DF")
            pdf.set_xy(x + 3, card_y + 3.5)
            self._set_text_color(pdf, BRAND_BLUE)
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(
                card_width - 6,
                6,
                str(value),
                new_x=XPos.LMARGIN,
                new_y=YPos.NEXT,
                align="C",
            )
            pdf.set_x(x + 3)
            self._set_text_color(pdf, MUTED)
            pdf.set_font("Helvetica", "B", 6.8)
            pdf.multi_cell(card_width - 6, 4, label.upper(), align="C")
        pdf.set_y(y + (card_height + gap) * 2)

    def _draw_signal_analytics(self, pdf: FPDF) -> None:
        self._section_title(pdf, "Signal Analytics")
        page_width = pdf.w - pdf.l_margin - pdf.r_margin
        gap = 5
        chart_width = (page_width - gap) / 2
        chart_height = 52
        y = pdf.get_y()

        self._draw_horizontal_bar_chart(
            pdf,
            title="Top Event Types",
            data=self.analytics.event_type_counts[:6],
            x=pdf.l_margin,
            y=y,
            width=chart_width,
            height=chart_height,
            color=BRAND_BLUE,
        )
        self._draw_horizontal_bar_chart(
            pdf,
            title="Threats by Marketplace",
            data=self.analytics.marketplace_counts[:6],
            x=pdf.l_margin + chart_width + gap,
            y=y,
            width=chart_width,
            height=chart_height,
            color=GREEN,
        )
        pdf.set_y(y + chart_height + 4)
        self._draw_insight_strip(pdf)

    def _draw_horizontal_bar_chart(
        self,
        pdf: FPDF,
        *,
        title: str,
        data: list[tuple[str, int]],
        x: float,
        y: float,
        width: float,
        height: float,
        color: str,
    ) -> None:
        self._set_fill_color(pdf, "FFFFFF")
        pdf.set_draw_color(*_pdf_rgb(BORDER))
        pdf.rect(x, y, width, height, "DF")
        pdf.set_xy(x + 3, y + 3)
        self._set_text_color(pdf, BRAND_NAVY)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(width - 6, 4, title)

        if not data:
            pdf.set_xy(x + 3, y + 22)
            self._set_text_color(pdf, MUTED)
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(width - 6, 4, "No signal data available")
            return

        max_value = max(value for _, value in data)
        label_width = width * 0.45
        bar_width = width - label_width - 12
        row_height = min(6.8, (height - 14) / max(len(data), 1))
        start_y = y + 12
        for index, (label, value) in enumerate(data):
            row_y = start_y + index * row_height
            pdf.set_xy(x + 3, row_y)
            self._set_text_color(pdf, INK)
            pdf.set_font("Helvetica", "", 6.7)
            pdf.cell(label_width, 3.8, _safe_pdf_text(label[:28]))
            self._set_fill_color(pdf, SURFACE)
            pdf.rect(x + 3 + label_width, row_y + 0.6, bar_width, 3.2, "F")
            self._set_fill_color(pdf, color)
            filled = bar_width * (value / max_value if max_value else 0)
            pdf.rect(x + 3 + label_width, row_y + 0.6, filled, 3.2, "F")
            pdf.set_xy(x + width - 8, row_y)
            self._set_text_color(pdf, MUTED)
            pdf.set_font("Helvetica", "B", 6.5)
            pdf.cell(5, 3.8, str(value), align="R")

    def _draw_insight_strip(self, pdf: FPDF) -> None:
        page_width = pdf.w - pdf.l_margin - pdf.r_margin
        insights = [
            ("Top Signal", self.analytics.top_event_type),
            ("Avg Signals / Threat", f"{self.analytics.average_signals_per_threat:.1f}"),
            ("Top Tracker", self.analytics.top_tracker),
        ]
        width = page_width / len(insights)
        y = pdf.get_y()
        self._set_fill_color(pdf, SURFACE)
        pdf.set_draw_color(*_pdf_rgb(BORDER))
        pdf.rect(pdf.l_margin, y, page_width, 18, "DF")
        for index, (label, value) in enumerate(insights):
            x = pdf.l_margin + index * width + 4
            pdf.set_xy(x, y + 4)
            self._set_text_color(pdf, MUTED)
            pdf.set_font("Helvetica", "B", 6.7)
            pdf.cell(width - 8, 3.5, label.upper())
            pdf.set_xy(x, y + 9)
            self._set_text_color(pdf, INK)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(width - 8, 4, _safe_pdf_text(value[:34]))
        pdf.set_y(y + 20)

    def _draw_ai_insights(self, pdf: FPDF) -> None:
        assert self.digest.insights is not None
        insights = self.digest.insights
        page_width = pdf.w - pdf.l_margin - pdf.r_margin

        self._section_title(pdf, "AI Insights")
        y = pdf.get_y()
        self._set_fill_color(pdf, SURFACE)
        pdf.set_draw_color(*_pdf_rgb(BORDER))
        content_height = 12 + len(insights.key_trends) * 5 + 14
        pdf.rect(pdf.l_margin, y, page_width, content_height, "DF")

        cursor_y = y + 4
        pdf.set_xy(pdf.l_margin + 4, cursor_y)
        self._set_text_color(pdf, BRAND_NAVY)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(page_width - 8, 4, "Executive Summary")
        cursor_y += 5
        pdf.set_x(pdf.l_margin + 4)
        self._set_text_color(pdf, INK)
        pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(page_width - 8, 4, _safe_pdf_text(insights.executive_summary))
        cursor_y = pdf.get_y() + 3

        pdf.set_xy(pdf.l_margin + 4, cursor_y)
        self._set_text_color(pdf, BRAND_NAVY)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(page_width - 8, 4, "Key Trends")
        cursor_y += 5
        for trend in insights.key_trends:
            pdf.set_x(pdf.l_margin + 4)
            self._set_text_color(pdf, INK)
            pdf.set_font("Helvetica", "", 7.5)
            pdf.multi_cell(page_width - 8, 4, _safe_pdf_text(f"• {trend}"))
            cursor_y = pdf.get_y() + 1
        cursor_y += 2

        pdf.set_xy(pdf.l_margin + 4, cursor_y)
        self._set_text_color(pdf, BRAND_NAVY)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(page_width - 8, 4, "Risk Assessment")
        cursor_y += 5
        pdf.set_x(pdf.l_margin + 4)
        self._set_text_color(pdf, INK)
        pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(page_width - 8, 4, _safe_pdf_text(insights.risk_assessment))

        pdf.set_y(y + content_height + 4)

    def _draw_trackers(self, pdf: FPDF) -> None:
        self._section_title(pdf, "Trackers Included")
        page_width = pdf.w - pdf.l_margin - pdf.r_margin
        name_width = page_width * 0.68
        type_width = page_width - name_width
        self._table_header(pdf, [("Tracker Name", name_width), ("Type", type_width)])
        for ref in self.digest.tracker_refs:
            y = pdf.get_y()
            row_height = 8
            pdf.set_draw_color(*_pdf_rgb(BORDER))
            pdf.rect(pdf.l_margin, y, name_width, row_height)
            pdf.rect(pdf.l_margin + name_width, y, type_width, row_height)
            pdf.set_xy(pdf.l_margin + 2, y + 2)
            pdf.set_font("Helvetica", "", 8.5)
            self._set_text_color(pdf, INK)
            pdf.cell(name_width - 4, 4, _safe_pdf_text(ref.tracker_name))
            pdf.set_xy(pdf.l_margin + name_width + 2, y + 2)
            self._set_text_color(pdf, BRAND_NAVY)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(type_width - 4, 4, _safe_pdf_text(ref.tracker_type.value))
            pdf.set_y(y + row_height)

    def _draw_threats(self, pdf: FPDF) -> None:
        self._section_title(pdf, f"Threat Analysis ({len(self.digest.threats)})")
        page_width = pdf.w - pdf.l_margin - pdf.r_margin
        if not self.digest.threats:
            self._set_text_color(pdf, MUTED)
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(
                0,
                7,
                "No threats detected during this period.",
                new_x=XPos.LMARGIN,
                new_y=YPos.NEXT,
            )
            return

        for i, threat in enumerate(self.digest.threats, 1):
            min_height = 32
            if pdf.get_y() + min_height > pdf.page_break_trigger:
                pdf.add_page()
            y = pdf.get_y()
            self._set_fill_color(pdf, WARN_FILL if i == 1 else "FFFFFF")
            pdf.set_draw_color(*_pdf_rgb(BORDER))
            pdf.rect(pdf.l_margin, y, page_width, min_height, "DF")
            pdf.set_xy(pdf.l_margin + 4, y + 4)
            self._set_text_color(pdf, BRAND_NAVY)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(
                page_width - 8,
                5,
                _safe_pdf_text(f"{i}. {threat.asin} | {threat.marketplace}"),
                new_x=XPos.LMARGIN,
                new_y=YPos.NEXT,
            )
            pdf.set_x(pdf.l_margin + 4)
            self._set_text_color(pdf, INK)
            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(page_width - 8, 5, _safe_pdf_text(f"Reason: {threat.reason}"))
            pdf.set_x(pdf.l_margin + 4)
            self._set_text_color(pdf, MUTED)
            pdf.set_font("Helvetica", "B", 7.5)
            pdf.multi_cell(
                page_width - 8,
                4.5,
                _safe_pdf_text(f"Event Types: {_join_values(threat.event_types)}"),
            )
            pdf.set_x(pdf.l_margin + 4)
            trackers = ", ".join(ref.tracker_name for ref in threat.tracker_refs)
            pdf.multi_cell(
                page_width - 8,
                4.5,
                _safe_pdf_text(f"Trackers: {trackers}"),
            )
            pdf.set_y(max(pdf.get_y() + 3, y + min_height + 3))

    def _table_header(self, pdf: FPDF, columns: list[tuple[str, float]]) -> None:
        y = pdf.get_y()
        x = pdf.l_margin
        self._set_fill_color(pdf, HEADER_FILL)
        pdf.set_draw_color(*_pdf_rgb(BORDER))
        pdf.set_font("Helvetica", "B", 8)
        self._set_text_color(pdf, BRAND_NAVY)
        for label, width in columns:
            pdf.rect(x, y, width, 7, "DF")
            pdf.set_xy(x + 2, y + 2)
            pdf.cell(width - 4, 3, label)
            x += width
        pdf.set_y(y + 7)


class ReportExcelGenerator:
    def __init__(self, digest: "WeeklyDigest") -> None:
        self.digest = digest
        self.analytics = _build_analytics(digest)

    def generate(self) -> bytes:
        wb = Workbook()
        wb.properties.title = "Weekly Digest Report"
        wb.properties.subject = self.digest.digest_code
        wb.properties.creator = "Market Tracker"

        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.sheet_properties.tabColor = BRAND_NAVY

        ws_summary["A1"] = "Weekly Digest Report"
        ws_summary["A1"].font = Font(bold=True, size=18, color="FFFFFF")
        ws_summary["A1"].fill = PatternFill("solid", fgColor=BRAND_NAVY)
        ws_summary["A1"].alignment = Alignment(vertical="center")
        ws_summary.merge_cells("A1:E1")
        ws_summary.row_dimensions[1].height = 28

        ws_summary["A3"] = "Week"
        ws_summary["B3"] = f"{self.digest.week_start} to {self.digest.week_end}"
        ws_summary["A4"] = "Code"
        ws_summary["B4"] = self.digest.digest_code
        ws_summary["A5"] = "Created"
        ws_summary["B5"] = self.digest.created_at.strftime("%Y-%m-%d %H:%M:%S")

        ws_summary["A7"] = "Metric"
        ws_summary["B7"] = "Count"
        self._style_header_row(ws_summary, 7, 1, 2)

        metrics = [
            ("New Entrants", self.digest.summary.new_entrant_count),
            ("Returning", self.digest.summary.returning_count),
            ("Top 10 Entry", self.digest.summary.top10_enter_count),
            ("Price Changes", self.digest.summary.price_change_count),
            ("Listing Changes", self.digest.summary.listing_change_count),
            ("Threats", len(self.digest.threats)),
            ("Event Signals", self.analytics.total_event_signals),
            ("Active Marketplaces", self.analytics.active_marketplaces),
            ("Avg Signals / Threat", self.analytics.average_signals_per_threat),
        ]
        for i, (label, value) in enumerate(metrics, start=8):
            ws_summary[f"A{i}"] = label
            ws_summary[f"B{i}"] = value
        self._style_summary_sheet(ws_summary)

        self._create_analytics_sheet(wb)

        if self.digest.insights:
            self._create_insights_sheet(wb)

        ws_trackers = wb.create_sheet("Trackers")
        ws_trackers.sheet_properties.tabColor = BRAND_BLUE
        ws_trackers["A1"] = "Tracker Name"
        ws_trackers["B1"] = "Type"
        self._style_header_row(ws_trackers, 1, 1, 2)
        for i, ref in enumerate(self.digest.tracker_refs, start=2):
            ws_trackers[f"A{i}"] = ref.tracker_name
            ws_trackers[f"B{i}"] = ref.tracker_type.value
        self._style_table_sheet(ws_trackers, widths={1: 42, 2: 18})

        ws_threats = wb.create_sheet("Threats")
        ws_threats.sheet_properties.tabColor = "DC2626"
        ws_threats["A1"] = "ASIN"
        ws_threats["B1"] = "Marketplace"
        ws_threats["C1"] = "Reason"
        ws_threats["D1"] = "Event Types"
        ws_threats["E1"] = "Trackers"
        self._style_header_row(ws_threats, 1, 1, 5)
        for i, threat in enumerate(self.digest.threats, start=2):
            ws_threats[f"A{i}"] = threat.asin
            ws_threats[f"B{i}"] = threat.marketplace
            ws_threats[f"C{i}"] = threat.reason
            ws_threats[f"D{i}"] = _join_values(threat.event_types)
            ws_threats[f"E{i}"] = ", ".join(ref.tracker_name for ref in threat.tracker_refs)
        self._style_table_sheet(
            ws_threats,
            widths={1: 16, 2: 16, 3: 58, 4: 52, 5: 46},
            wrap_columns={3, 4, 5},
        )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _create_analytics_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Analytics")
        ws.sheet_properties.tabColor = AMBER
        ws["A1"] = "Signal Analytics"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=BRAND_NAVY)
        ws.merge_cells("A1:H1")
        ws.row_dimensions[1].height = 26

        self._write_count_table(
            ws,
            title="Event Type Distribution",
            start_row=3,
            start_col=1,
            headers=("Event Type", "Count"),
            rows=self.analytics.event_type_counts,
        )
        self._write_count_table(
            ws,
            title="Marketplace Mix",
            start_row=3,
            start_col=4,
            headers=("Marketplace", "Threats"),
            rows=self.analytics.marketplace_counts,
        )
        self._write_count_table(
            ws,
            title="Tracker Coverage",
            start_row=3,
            start_col=7,
            headers=("Tracker", "Threats"),
            rows=self.analytics.tracker_threat_counts,
        )

        self._add_bar_chart(
            ws,
            title="Top Event Types",
            data_start_row=5,
            data_end_row=max(5, 4 + len(self.analytics.event_type_counts)),
            label_col=1,
            value_col=2,
            anchor="A16",
        )
        self._add_pie_chart(
            ws,
            title="Threats by Marketplace",
            data_start_row=5,
            data_end_row=max(5, 4 + len(self.analytics.marketplace_counts)),
            label_col=4,
            value_col=5,
            anchor="D16",
        )
        self._add_bar_chart(
            ws,
            title="Tracker Threat Coverage",
            data_start_row=5,
            data_end_row=max(5, 4 + len(self.analytics.tracker_threat_counts)),
            label_col=7,
            value_col=8,
            anchor="G16",
        )

        widths = {1: 30, 2: 12, 4: 20, 5: 12, 7: 34, 8: 12}
        for index, width in widths.items():
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = "A5"

    def _create_insights_sheet(self, wb: Workbook) -> None:
        assert self.digest.insights is not None
        insights = self.digest.insights

        ws = wb.create_sheet("AI Insights")
        ws.sheet_properties.tabColor = "7C3AED"

        ws["A1"] = "AI Insights"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=BRAND_NAVY)
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 26

        ws["A3"] = "Executive Summary"
        ws["A3"].font = Font(bold=True, color=BRAND_NAVY)
        ws["A4"] = insights.executive_summary
        ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A4:F4")
        ws.row_dimensions[4].height = 60

        current_row = 6
        ws.cell(row=current_row, column=1, value="Key Trends")
        ws.cell(row=current_row, column=1).font = Font(bold=True, color=BRAND_NAVY)
        current_row += 1
        for i, trend in enumerate(insights.key_trends, start=1):
            ws.cell(row=current_row, column=1, value=f"{i}.")
            ws.cell(row=current_row, column=1).font = Font(color=MUTED)
            ws.cell(row=current_row, column=2, value=trend)
            ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
            ws.merge_cells(
                start_row=current_row, start_column=2,
                end_row=current_row, end_column=6,
            )
            current_row += 1

        current_row += 1
        ws.cell(row=current_row, column=1, value="Risk Assessment")
        ws.cell(row=current_row, column=1).font = Font(bold=True, color=BRAND_NAVY)
        current_row += 1
        ws.cell(row=current_row, column=1, value=insights.risk_assessment)
        ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=6,
        )
        ws.row_dimensions[current_row].height = 45

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 60
        for col in "CDEF":
            ws.column_dimensions[col].width = 16

    def _write_count_table(
        self,
        ws: Worksheet,
        *,
        title: str,
        start_row: int,
        start_col: int,
        headers: tuple[str, str],
        rows: list[tuple[str, int]],
    ) -> None:
        ws.cell(row=start_row, column=start_col, value=title)
        ws.cell(row=start_row, column=start_col).font = Font(bold=True, color=BRAND_NAVY)
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=start_col + 1,
        )

        header_row = start_row + 1
        ws.cell(row=header_row, column=start_col, value=headers[0])
        ws.cell(row=header_row, column=start_col + 1, value=headers[1])
        self._style_header_row(ws, header_row, start_col, start_col + 1)

        data_rows = rows or [("No data", 0)]
        for offset, (label, value) in enumerate(data_rows, start=1):
            row = header_row + offset
            ws.cell(row=row, column=start_col, value=label)
            ws.cell(row=row, column=start_col + 1, value=value)
        self._apply_grid(
            ws,
            min_row=header_row,
            max_row=header_row + len(data_rows),
            min_col=start_col,
            max_col=start_col + 1,
        )

    def _add_bar_chart(
        self,
        ws: Worksheet,
        *,
        title: str,
        data_start_row: int,
        data_end_row: int,
        label_col: int,
        value_col: int,
        anchor: str,
    ) -> None:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = title
        chart.x_axis.title = "Count"
        chart.height = 8
        chart.width = 11
        data = Reference(
            ws,
            min_col=value_col,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        labels = Reference(
            ws,
            min_col=label_col,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        chart.legend = None
        ws.add_chart(chart, anchor)

    def _add_pie_chart(
        self,
        ws: Worksheet,
        *,
        title: str,
        data_start_row: int,
        data_end_row: int,
        label_col: int,
        value_col: int,
        anchor: str,
    ) -> None:
        chart = PieChart()
        chart.title = title
        chart.height = 8
        chart.width = 10
        data = Reference(
            ws,
            min_col=value_col,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        labels = Reference(
            ws,
            min_col=label_col,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        ws.add_chart(chart, anchor)

    def _style_summary_sheet(self, ws: Worksheet) -> None:
        widths = {1: 22, 2: 30, 3: 16, 4: 16, 5: 16}
        for index, width in widths.items():
            ws.column_dimensions[get_column_letter(index)].width = width
        for row in range(3, 6):
            ws[f"A{row}"].font = Font(bold=True, color=MUTED)
            ws[f"B{row}"].font = Font(color=INK)
        for row in range(8, 17):
            ws[f"A{row}"].fill = PatternFill("solid", fgColor=SURFACE)
            ws[f"B{row}"].fill = PatternFill("solid", fgColor=SURFACE)
            ws[f"B{row}"].font = Font(bold=True, color=BRAND_BLUE)
            ws[f"B{row}"].alignment = Alignment(horizontal="right")
        self._apply_grid(ws, min_row=3, max_row=16, min_col=1, max_col=2)
        ws.freeze_panes = "A7"

    def _style_table_sheet(
        self,
        ws: Worksheet,
        *,
        widths: dict[int, int],
        wrap_columns: set[int] | None = None,
    ) -> None:
        wrap_columns = wrap_columns or set()
        for index, width in widths.items():
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        self._apply_grid(
            ws,
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        )
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=cell.column in wrap_columns,
                )

    def _style_header_row(
        self, ws: Worksheet, row: int, min_col: int, max_col: int
    ) -> None:
        for column in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=column)
            cell.font = Font(bold=True, color=BRAND_NAVY)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22

    def _apply_grid(
        self,
        ws: Worksheet,
        *,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> None:
        thin = Side(style="thin", color=BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for cell in row:
                cell.border = border
                alignment = copy(cell.alignment)
                alignment.vertical = "top"
                cell.alignment = alignment
